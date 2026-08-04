"""Score SK hynix Q2 2026 actuals against the immutable July 10 workbook."""

from __future__ import annotations

import argparse
import hashlib
import logging
import os
import tempfile
from dataclasses import dataclass
from pathlib import Path

import yaml
from openpyxl import load_workbook

logger = logging.getLogger(__name__)

REPO_ROOT = Path(__file__).resolve().parent.parent
ANCHOR_PATH = REPO_ROOT / "reports" / "sk_hynix_20260710.xlsx"
ACTUAL_PATH = REPO_ROOT / "inputs" / "sk_hynix_q2_2026_actual.yaml"
SCORECARD_PATH = REPO_ROOT / "reports" / "sk_hynix_q2_2026_scorecard.md"
ANCHOR_SHA256 = "e2ed59edb4abac6709b6393f34892dd641d0ff6bc55431da3d4cb53b86166f95"
PROFILE_SHARES = 705_656_476.0
START_MARKER = "<!-- SK_HYNIX_Q2_2026_T1_START -->"
END_MARKER = "<!-- SK_HYNIX_Q2_2026_T1_END -->"


@dataclass(frozen=True)
class Point:
    revenue: float
    operating_profit: float
    net_income: float
    eps: float


@dataclass(frozen=True)
class Attribution:
    revenue: float
    operating_margin: float
    tax_finance: float
    shares: float
    total: float
    residual: float
    revenue_model_eps: float
    operating_margin_model_eps: float
    tax_finance_model_eps: float
    shares_model_eps: float
    total_model_eps: float
    residual_model_eps: float
    shares_lever_degenerate: bool
    comparable_to_backtest_five_lever: bool


def sha256(path: Path) -> str:
    """Return the SHA-256 digest for a file."""
    return hashlib.sha256(path.read_bytes()).hexdigest()


def load_anchor(path: Path = ANCHOR_PATH) -> Point:
    """Load the weighted 2026Q2 point from the read-only workbook."""
    workbook = load_workbook(path, data_only=True, read_only=True)
    sheet = workbook["forecast"]
    if sheet.cell(2, 1).value != "2026Q2":
        raise ValueError("anchor forecast row 2 is not 2026Q2")
    return Point(*(float(sheet.cell(2, column).value) for column in range(2, 6)))


def load_actual(path: Path = ACTUAL_PATH) -> tuple[Point, dict]:
    """Load the actual point and retain metadata used by the report."""
    payload = yaml.safe_load(path.read_text(encoding="utf-8"))
    actuals = payload["actuals"]
    required = ("revenue_krw_bn", "operating_profit_krw_bn", "net_income_krw_bn", "eps_krw")
    missing = [name for name in required if actuals[name]["value"] is None]
    if missing:
        raise ValueError(f"missing required actuals: {', '.join(missing)}")
    return (
        Point(*(float(actuals[name]["value"]) for name in required)),
        payload,
    )


def attribute_eps_error(model: Point, actual: Point, payload: dict) -> Attribution:
    """Return a telescoping four-lever EPS attribution.

    The preliminary release omits gross profit, so the memory path's gross-margin
    and GP-to-OP levers cannot be identified. This reduced bridge uses operating
    margin instead and is not directly comparable with the backtest five-lever
    waterfall.
    """
    for label, value in vars(model).items():
        if value == 0:
            raise ValueError(f"model {label} must be non-zero")
    for label, value in vars(actual).items():
        if value == 0:
            raise ValueError(f"actual {label} must be non-zero")

    actual_share_value = payload["actuals"]["weighted_avg_shares"]["value"]
    actual_shares = PROFILE_SHARES if actual_share_value is None else float(actual_share_value)
    derived_eps = actual.net_income * 1_000_000_000.0 / actual_shares
    if abs(derived_eps - actual.eps) > 1e-9 * abs(actual.eps):
        raise ValueError("actual EPS is inconsistent with net income and shares")

    op_margin_model = model.operating_profit / model.revenue
    op_margin_actual = actual.operating_profit / actual.revenue
    ni_to_op_model = model.net_income / model.operating_profit
    ni_to_op_actual = actual.net_income / actual.operating_profit
    inverse_shares_model = 1_000_000_000.0 / PROFILE_SHARES
    inverse_shares_actual = 1_000_000_000.0 / actual_shares

    def eps(revenue: float, op_margin: float, ni_to_op: float, inverse_shares: float) -> float:
        return revenue * op_margin * ni_to_op * inverse_shares

    def error(value: float) -> float:
        return (value - actual.eps) / actual.eps

    e0 = eps(model.revenue, op_margin_model, ni_to_op_model, inverse_shares_model)
    e1 = eps(actual.revenue, op_margin_model, ni_to_op_model, inverse_shares_model)
    e2 = eps(actual.revenue, op_margin_actual, ni_to_op_model, inverse_shares_model)
    e3 = eps(actual.revenue, op_margin_actual, ni_to_op_actual, inverse_shares_model)
    e4 = eps(actual.revenue, op_margin_actual, ni_to_op_actual, inverse_shares_actual)

    contributions = (
        error(e0) - error(e1),
        error(e1) - error(e2),
        error(e2) - error(e3),
        error(e3) - error(e4),
    )
    total = error(e0)
    residual = total - sum(contributions)
    model_scale = actual.eps / model.eps
    model_contributions = tuple(value * model_scale for value in contributions)
    total_model_eps = (model.eps - actual.eps) / model.eps
    residual_model_eps = total_model_eps - sum(model_contributions)
    degenerate = actual_share_value is None and payload["eps_derived"]
    return Attribution(
        *contributions,
        total=total,
        residual=residual,
        revenue_model_eps=model_contributions[0],
        operating_margin_model_eps=model_contributions[1],
        tax_finance_model_eps=model_contributions[2],
        shares_model_eps=model_contributions[3],
        total_model_eps=total_model_eps,
        residual_model_eps=residual_model_eps,
        shares_lever_degenerate=degenerate,
        comparable_to_backtest_five_lever=False,
    )


def render(model: Point, actual: Point, attribution: Attribution) -> str:
    """Render the idempotent scorecard block."""
    def pct(value: float) -> str:
        return f"{value * 100:+.4f}%"

    return "\n".join(
        [
            START_MARKER,
            "## 9. T1 재현 채점 (Codex)",
            "",
            "> **사후 귀인 — 예측 신호 아님.** T0b에서 7/10 앵커가 원시 셀 기준 relative error 0으로 재현됐고, 9Q SHA도 canonical MATCH였다.",
            "",
            "| 지표 | 모델 weighted | 실측 | 오차 |",
            "|---|---:|---:|---:|",
            f"| 매출 (KRW bn) | {model.revenue:,.2f} | {actual.revenue:,.1f} | {pct((model.revenue - actual.revenue) / actual.revenue)} |",
            f"| 영업이익 (KRW bn) | {model.operating_profit:,.2f} | {actual.operating_profit:,.1f} | {pct((model.operating_profit - actual.operating_profit) / actual.operating_profit)} |",
            f"| 순이익 (KRW bn) | {model.net_income:,.2f} | {actual.net_income:,.1f} | {pct((model.net_income - actual.net_income) / actual.net_income)} |",
            f"| EPS (KRW) | {model.eps:,.2f} | {actual.eps:,.2f} | {pct(attribution.total)} |",
            "",
            "### 4레버 축약 EPS 오차 귀인",
            "",
            "> **잠정자료 제약에 따른 축약 — 백테스트 5레버 워터폴과 직접 비교 불가. 반기보고서 GP 확인 시 5레버로 복원한다.**",
            "",
            "> **별도 비교 제약 — 정규화 압축.** 현 규약은 실측 EPS를 분모로 쓴다. 이번 실측 EPS는 일회성 below-OP 이익으로 모델의 1.89배이므로 모든 레버가 일률적으로 약 0.53배 압축된다. 축약 여부와 별개의 문제이며, 현 규약 열을 과거 분기의 레버 크기와 직접 비교하면 안 된다.",
            "",
            "| 레버 | ÷실측EPS (현 규약) | ÷모델EPS (비교용) |",
            "|---|---:|---:|",
            f"| 매출 | {pct(attribution.revenue)} | {pct(attribution.revenue_model_eps)} |",
            f"| OP마진 | {pct(attribution.operating_margin)} | {pct(attribution.operating_margin_model_eps)} |",
            f"| 세금·금융 | {pct(attribution.tax_finance)} | {pct(attribution.tax_finance_model_eps)} |",
            f"| 주식수 | {pct(attribution.shares)} | {pct(attribution.shares_model_eps)} |",
            f"| 합계 | {pct(attribution.total)} | {pct(attribution.total_model_eps)} |",
            f"| 잔차 | {attribution.residual:+.3e} | {attribution.residual_model_eps:+.3e} |",
            "",
            "- `÷실측EPS`는 `(모델 EPS − 실측 EPS) / 실측 EPS`로 정의한 표준 오차율이다.",
            "- `÷모델EPS`는 동일한 절대 레버 금액을 모델 EPS로 나눈 분기 간 비교용 값이다. 합계는 `(모델 EPS − 실측 EPS) / 모델 EPS`이며 두 열을 섞어 인용하면 안 된다.",
            "",
            "- `shares_lever_degenerate: true`: 실제 EPS가 모델과 동일한 705,656,476주로 파생됐으므로 주식수 레버 0은 정확도 증거가 아니다. 이 분기의 EPS 귀인은 실질적으로 NI 귀인이다.",
            "- 매출 레버/매출 오차 비율 0.545는 실측 OP마진 앵커 때문이 아니다. 순차 치환식상 비율은 `(model EPS / actual EPS) × (actual revenue / model revenue)`이며, 이번 분기의 대형 below-OP 이익이 이 비율을 압축했다.",
            "- 세전이익·법인세·영업외손익·실제 가중평균주식수는 미확정이며 추정치를 actual에 혼입하지 않았다.",
            "- DRAM ASP +30%가 HBM 포함 blended인지 conventional 한정인지 공개 문구만으로 확정되지 않아 정의 미확정으로 기록했다.",
            "- 키옥시아 관련 이익의 P&L/OCI, 지분매각/CB 평가 구분과 잔여 지분은 1차 출처에서 확정되지 않았다. H2는 반기보고서 전까지 검증되지 않은 가설이다.",
            "",
            "### 프로비넌스 등급",
            "",
            "**5종목과 동등.** 근거는 §8의 9Q canonical SHA MATCH, live forward 원시 셀 relative error 0, 계산 경로의 docstring-only diff다.",
            END_MARKER,
            "",
        ]
    )


def append_block(path: Path, block: str) -> None:
    """Append or replace the marked block atomically."""
    original = path.read_text(encoding="utf-8")
    if START_MARKER in original:
        start = original.index(START_MARKER)
        end = original.index(END_MARKER, start) + len(END_MARKER)
        updated = original[:start] + block.rstrip() + original[end:]
    else:
        updated = original.rstrip() + "\n\n---\n\n" + block
    with tempfile.NamedTemporaryFile(
        mode="w", encoding="utf-8", newline="\n", delete=False, dir=path.parent
    ) as handle:
        handle.write(updated)
        temporary = Path(handle.name)
    try:
        os.replace(temporary, path)
    finally:
        if temporary.exists():
            temporary.unlink()


def main() -> None:
    """Run scoring and optionally update the scorecard."""
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--append", action="store_true")
    args = parser.parse_args()
    logging.basicConfig(level=logging.INFO, format="%(message)s")

    before = sha256(ANCHOR_PATH)
    if before != ANCHOR_SHA256:
        raise RuntimeError(f"anchor SHA mismatch: {before}")
    model = load_anchor()
    actual, payload = load_actual()
    attribution = attribute_eps_error(model, actual, payload)
    if abs(attribution.residual) >= 1e-9:
        raise RuntimeError(f"attribution residual too large: {attribution.residual}")
    if abs(attribution.residual_model_eps) >= 1e-9:
        raise RuntimeError(
            "model-EPS-normalized attribution residual too large: "
            f"{attribution.residual_model_eps}"
        )
    block = render(model, actual, attribution)
    logger.info("%s", block)
    if args.append:
        append_block(SCORECARD_PATH, block)
        logger.info("updated %s", SCORECARD_PATH)
    if sha256(ANCHOR_PATH) != before:
        raise RuntimeError("anchor workbook changed during scoring")


if __name__ == "__main__":
    main()
