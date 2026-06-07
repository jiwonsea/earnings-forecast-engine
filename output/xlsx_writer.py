"""xlsx writer — 3 sheets: forecast, scenario, backtest.

Raw numeric verification artifact. Not used for presentation — that role
belongs to HTML / MD.
"""

from __future__ import annotations

from pathlib import Path

from schemas.models import BacktestResult, ScenarioTree


def write_xlsx(
    out_path: Path,
    tree: ScenarioTree,
    backtest: BacktestResult,
) -> Path:
    """Write a 3-sheet xlsx via openpyxl.

    Sheet 1 ("forecast"): weighted quarterly + annual P&L.
    Sheet 2 ("scenarios"): bear / base / bull side-by-side per quarter.
    Sheet 3 ("backtest"): per-quarter detail + summary stats.

    Args:
        out_path: Target file path (.xlsx).
        tree, backtest: Inputs.

    Returns:
        out_path on success.

    Raises:
        NotImplementedError: Codex implementation.
    """
    from openpyxl import Workbook

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "forecast"
    ws.append(["quarter", "revenue", "op", "ni", "eps"])
    for q in tree.weighted_quarterly:
        ws.append([q.quarter_label, q.revenue_total, q.operating_profit, q.net_profit, q.eps_basic])
    ws.append([])
    ws.append(["year", "revenue", "op", "ni", "eps"])
    for a in tree.weighted_annual:
        ws.append([a.fiscal_year, a.revenue_total, a.operating_profit, a.net_profit, a.eps_basic])

    ws = wb.create_sheet("scenarios")
    ws.append(["quarter", "bear", "base", "bull", "weighted"])
    for idx, q in enumerate(tree.base.quarterly):
        ws.append(
            [
                q.quarter_label,
                tree.bear.quarterly[idx].revenue_total,
                tree.base.quarterly[idx].revenue_total,
                tree.bull.quarterly[idx].revenue_total,
                tree.weighted_quarterly[idx].revenue_total,
            ]
        )

    ws = wb.create_sheet("backtest")
    ws.append(["quarter", "actual_revenue", "model_revenue", "revenue_error_pct", "actual_eps", "model_eps", "eps_error_pct", "direction_match"])
    for q in backtest.quarters:
        ws.append([q.quarter_label, q.actual_revenue, q.model_revenue, q.revenue_error_pct, q.actual_eps, q.model_eps, q.eps_error_pct, q.direction_match])
    ws.append([])
    ws.append(["revenue_mape", backtest.revenue_mape])
    ws.append(["eps_mape", backtest.eps_mape])
    ws.append(["hit_ratio_direction", backtest.hit_ratio_direction])
    wb.save(out_path)
    return out_path
